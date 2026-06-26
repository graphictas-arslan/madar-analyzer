from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

def generate_excel(posts, title="پست‌ها"):
    """
    تولید فایل اکسل از لیست پست‌ها
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    headers = ['شناسه', 'کانال', 'نوع', 'متن', 'کپشن', 'تاریخ انتشار', 'وضعیت', 'امتیاز', 'بازدید', 'لایک', 'کامنت']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for post in posts:
        from models import Channel
        channel = Channel.query.get(post.channel_id)
        channel_name = channel.channel_name if channel else '-'
        
        ws.append([
            post.id,
            channel_name,
            post.post_type,
            post.text or '',
            post.caption or '',
            post.publish_date.strftime('%Y-%m-%d %H:%M') if post.publish_date else '',
            post.status,
            post.score or 0,
            post.views or 0,
            post.likes or 0,
            post.comments or 0
        ])
    
    column_widths = [10, 20, 10, 40, 30, 20, 15, 10, 10, 10, 10]
    for i, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
    
    ws.sheet_view.rightToLeft = True
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
